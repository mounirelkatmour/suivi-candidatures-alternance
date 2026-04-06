import customtkinter as ctk
from tkinter import messagebox, ttk
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Config ──────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

FICHIER_EXCEL = "Suivi Candidatures Alternance.xlsx"
COLONNES = ["Entreprise", "Contact", "Poste", "Canal", "Date envoi", "Statut", "Notes", "Ville", "Priorité"]
CANAUX = ["LinkedIn", "Email", "Site web"]
STATUTS = ["En attente", "✅ Réponse", "❌", "Relancé", "Entretien"]
PRIORITES = ["⭐", "⭐⭐", "⭐⭐⭐"]

# ── Excel helpers ────────────────────────────────────────────────────────────
def init_excel():
    if not os.path.exists(FICHIER_EXCEL):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Suivi Candidatures"
        for col, h in enumerate(COLONNES, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color="1F4E79")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(FICHIER_EXCEL)

def lire_donnees():
    wb = openpyxl.load_workbook(FICHIER_EXCEL)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(list(row))
    return rows

def ecrire_donnees(rows):
    wb = openpyxl.load_workbook(FICHIER_EXCEL)
    ws = wb.active
    # Clear existing data
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    status_colors = {
        "En attente": "FFF2CC",
        "✅ Réponse": "E2EFDA",
        "❌": "FFE0E0",
        "Relancé": "FCE4D6",
        "Entretien": "D6E4F0",
    }

    for r_idx, row in enumerate(rows, 2):
        bg = "F2F2F2" if r_idx % 2 == 0 else "FFFFFF"
        status = row[5] if len(row) > 5 else ""
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            thin = Side(style="thin", color="CCCCCC")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if c_idx == 6:
                cell.fill = PatternFill("solid", start_color=status_colors.get(status, bg))
                cell.font = Font(name="Arial", size=10, bold=True)
            else:
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(name="Arial", size=10)

    widths = [22, 22, 25, 12, 14, 14, 40, 12, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(FICHIER_EXCEL)

# ── Formulaire (ajout / modification) ───────────────────────────────────────
class Formulaire(ctk.CTkToplevel):
    def __init__(self, parent, callback, donnees=None):
        super().__init__(parent)
        self.title("Modifier" if donnees else "Ajouter une candidature")
        self.geometry("500x600")
        self.resizable(False, False)
        self.grab_set()
        self.callback = callback
        self.donnees = donnees

        labels = COLONNES
        self.champs = {}

        for i, label in enumerate(labels):
            ctk.CTkLabel(self, text=label, anchor="w").grid(row=i, column=0, padx=20, pady=6, sticky="w")

            if label == "Canal":
                widget = ctk.CTkComboBox(self, values=CANAUX, width=280)
                if donnees and donnees[i]:
                    widget.set(donnees[i])
                else:
                    widget.set(CANAUX[0])
            elif label == "Statut":
                widget = ctk.CTkComboBox(self, values=STATUTS, width=280)
                if donnees and donnees[i]:
                    widget.set(donnees[i])
                else:
                    widget.set(STATUTS[0])
            elif label == "Priorité":
                widget = ctk.CTkComboBox(self, values=PRIORITES, width=280)
                if donnees and donnees[i]:
                    widget.set(donnees[i])
                else:
                    widget.set(PRIORITES[1])
            elif label == "Notes":
                widget = ctk.CTkTextbox(self, width=280, height=80)
                if donnees and donnees[i]:
                    widget.insert("0.0", donnees[i])
            else:
                widget = ctk.CTkEntry(self, width=280)
                if donnees and donnees[i]:
                    widget.insert(0, donnees[i])

            widget.grid(row=i, column=1, padx=20, pady=6, sticky="w")
            self.champs[label] = widget

        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.grid(row=len(labels), column=0, columnspan=2, pady=20)

        ctk.CTkButton(btn_frame, text="💾 Enregistrer", command=self.enregistrer, width=150).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Annuler", command=self.destroy, width=150, fg_color="gray").pack(side="left", padx=10)

    def enregistrer(self):
        row = []
        for label in COLONNES:
            widget = self.champs[label]
            if isinstance(widget, ctk.CTkTextbox):
                val = widget.get("0.0", "end").strip()
            elif isinstance(widget, ctk.CTkComboBox):
                val = widget.get()
            else:
                val = widget.get().strip()
            row.append(val if val else None)
        self.callback(row)
        self.destroy()

# ── Fenêtre principale ───────────────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("📋 Suivi Candidatures Alternance")
        self.geometry("1300x700")
        init_excel()
        self.build_ui()
        self.charger()

    def build_ui(self):
        # Header
        header = ctk.CTkFrame(self, height=60, corner_radius=0)
        header.pack(fill="x")
        ctk.CTkLabel(header, text="📋 Suivi Candidatures Alternance",
                     font=ctk.CTkFont(size=18, weight="bold")).pack(side="left", padx=20, pady=10)
        ctk.CTkButton(header, text="➕ Ajouter", command=self.ajouter, width=130).pack(side="right", padx=10, pady=10)
        ctk.CTkButton(header, text="🔄 Actualiser", command=self.charger, width=130, fg_color="gray").pack(side="right", padx=5, pady=10)

        # Stats
        self.stats_frame = ctk.CTkFrame(self, height=40, corner_radius=0, fg_color="#1a1a2e")
        self.stats_frame.pack(fill="x")
        self.lbl_stats = ctk.CTkLabel(self.stats_frame, text="", font=ctk.CTkFont(size=12))
        self.lbl_stats.pack(side="left", padx=20, pady=8)

        # Tableau
        table_frame = ctk.CTkFrame(self)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", background="#2b2b2b", foreground="white",
                        fieldbackground="#2b2b2b", rowheight=30, font=("Arial", 10))
        style.configure("Treeview.Heading", background="#1F4E79", foreground="white",
                        font=("Arial", 10, "bold"))
        style.map("Treeview", background=[("selected", "#1F4E79")])

        cols = COLONNES
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")

        widths_px = [130, 140, 150, 80, 100, 100, 200, 80, 70]
        for col, w in zip(cols, widths_px):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="w")

        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Boutons modifier/supprimer
        btn_frame = ctk.CTkFrame(self, height=50, corner_radius=0, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkButton(btn_frame, text="✏️ Modifier", command=self.modifier, width=150).pack(side="left", padx=5)
        ctk.CTkButton(btn_frame, text="🗑️ Supprimer", command=self.supprimer, width=150, fg_color="#c0392b").pack(side="left", padx=5)

    def charger(self):
        self.rows = lire_donnees()
        for item in self.tree.get_children():
            self.tree.delete(item)

        status_tags = {
            "En attente": "attente",
            "✅ Réponse": "reponse",
            "❌": "refus",
            "Relancé": "relance",
            "Entretien": "entretien",
        }

        self.tree.tag_configure("attente", background="#3d3d00")
        self.tree.tag_configure("reponse", background="#1a3300")
        self.tree.tag_configure("refus", background="#3d0000")
        self.tree.tag_configure("relance", background="#3d1a00")
        self.tree.tag_configure("entretien", background="#001a3d")

        for i, row in enumerate(self.rows):
            padded = (row + [None] * 9)[:9]
            status = padded[5] or ""
            tag = status_tags.get(status, "")
            self.tree.insert("", "end", iid=str(i), values=padded, tags=(tag,))

        # Stats
        total = len(self.rows)
        en_attente = sum(1 for r in self.rows if len(r) > 5 and r[5] == "En attente")
        reponses = sum(1 for r in self.rows if len(r) > 5 and r[5] == "✅ Réponse")
        refus = sum(1 for r in self.rows if len(r) > 5 and r[5] == "❌")
        self.lbl_stats.configure(
            text=f"Total : {total}  |  En attente : {en_attente}  |  Réponses : {reponses}  |  Refus : {refus}"
        )

    def get_selection(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Attention", "Veuillez sélectionner une ligne.")
            return None
        return int(sel[0])

    def ajouter(self):
        def callback(row):
            self.rows.append(row)
            ecrire_donnees(self.rows)
            self.charger()
        Formulaire(self, callback)

    def modifier(self):
        idx = self.get_selection()
        if idx is None:
            return
        def callback(row):
            self.rows[idx] = row
            ecrire_donnees(self.rows)
            self.charger()
        Formulaire(self, callback, donnees=self.rows[idx])

    def supprimer(self):
        idx = self.get_selection()
        if idx is None:
            return
        entreprise = self.rows[idx][0] if self.rows[idx] else "cette entrée"
        if messagebox.askyesno("Confirmer", f"Supprimer la candidature chez {entreprise} ?"):
            self.rows.pop(idx)
            ecrire_donnees(self.rows)
            self.charger()

if __name__ == "__main__":
    app = App()
    app.mainloop()
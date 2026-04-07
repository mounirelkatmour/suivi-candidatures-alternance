import customtkinter as ctk
from tkinter import messagebox, ttk, filedialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import json
import csv
from datetime import date, datetime, timedelta

# ── Config ───────────────────────────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

UI = {
    "bg": "#0b1220",
    "surface": "#111b2f",
    "surface_alt": "#16243d",
    "card": "#1b2b46",
    "card_hover": "#223757",
    "text": "#f4f7ff",
    "muted": "#9eb0cc",
    "line": "#2a3f63",
    "primary": "#2f7df6",
    "primary_hover": "#4e95ff",
    "danger": "#dd4b39",
    "danger_hover": "#f06252",
    "success": "#00a872",
}

TITLE_FONT = ("Segoe UI Semibold", 22, "bold")
H2_FONT = ("Segoe UI Semibold", 16, "bold")
BODY_FONT = ("Segoe UI", 12)
SMALL_FONT = ("Segoe UI", 11)

if getattr(sys, "frozen", False):
    # PyInstaller exe: keep user data next to the executable.
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    # Source mode: keep data in the project root.
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DATA_DIR = os.path.join(BASE_DIR, "data")
SUIVIS_JSON = os.path.join(DATA_DIR, "suivis.json")
SETTINGS_JSON = os.path.join(DATA_DIR, "settings.json")

COLONNES = ["Entreprise", "Contact", "Poste", "Canal", "Date envoi", "Statut", "Notes", "Ville", "Priorité"]
CANAUX = ["LinkedIn", "Email", "Site web"]
STATUTS = ["En attente", "✅ Réponse", "❌", "Relancé", "Entretien"]
PRIORITES = ["⭐", "⭐⭐", "⭐⭐⭐"]

STATUS_COLORS_EXCEL = {
    "En attente": "FFF2CC",
    "✅ Réponse": "E2EFDA",
    "❌": "FFE0E0",
    "Relancé": "FCE4D6",
    "Entretien": "D6E4F0",
}

STATUS_TAGS_UI = {
    "En attente": ("attente", "#3d3d00"),
    "✅ Réponse": ("reponse", "#1a3300"),
    "❌": ("refus", "#3d0000"),
    "Relancé": ("relance", "#3d1a00"),
    "Entretien": ("entretien", "#001a3d"),
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)

def load_suivis():
    ensure_data_dir()
    if not os.path.exists(SUIVIS_JSON):
        return []
    with open(SUIVIS_JSON, "r", encoding="utf-8") as f:
        return json.load(f)

def save_suivis(suivis):
    ensure_data_dir()
    with open(SUIVIS_JSON, "w", encoding="utf-8") as f:
        json.dump(suivis, f, ensure_ascii=False, indent=2)

def load_settings():
    ensure_data_dir()
    if not os.path.exists(SETTINGS_JSON):
        return {"weekly_goal": 15}
    with open(SETTINGS_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    if "weekly_goal" not in data:
        data["weekly_goal"] = 15
    return data

def save_settings(settings):
    ensure_data_dir()
    with open(SETTINGS_JSON, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)

def safe_filename(text):
    return "".join(c for c in text if c.isalnum() or c in " _-").strip() or "export"

def darken_hex(hex_color, factor=0.15):
    """Return a slightly darker color from a #RRGGBB input."""
    color = (hex_color or "").strip()
    if not color.startswith("#") or len(color) != 7:
        return hex_color
    try:
        r = int(color[1:3], 16)
        g = int(color[3:5], 16)
        b = int(color[5:7], 16)
    except ValueError:
        return hex_color

    scale = max(0.0, min(1.0, 1.0 - factor))
    r = int(r * scale)
    g = int(g * scale)
    b = int(b * scale)
    return f"#{r:02x}{g:02x}{b:02x}"

def parse_date_fr(value):
    if not value:
        return None
    try:
        return datetime.strptime(value.strip(), "%d/%m/%Y").date()
    except (ValueError, TypeError):
        return None

def sparkline(values):
    if not values:
        return ""
    ticks = "▁▂▃▄▅▆▇█"
    v_min = min(values)
    v_max = max(values)
    if v_min == v_max:
        return ticks[0] * len(values)
    return "".join(ticks[int((v - v_min) / (v_max - v_min) * (len(ticks) - 1))] for v in values)

def excel_path(nom):
    safe = "".join(c for c in nom if c.isalnum() or c in " _-").strip()
    return os.path.join(DATA_DIR, f"{safe}.xlsx")

def init_excel(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suivi Candidatures"
    for col, h in enumerate(COLONNES, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(path)

def lire_donnees(path):
    if not os.path.exists(path):
        return []
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append([str(c) if c is not None else "" for c in row])
    return rows

def ecrire_donnees(path, rows):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    for r_idx, row in enumerate(rows, 2):
        bg = "F2F2F2" if r_idx % 2 == 0 else "FFFFFF"
        status = row[5] if len(row) > 5 else ""
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val or None)
            thin = Side(style="thin", color="CCCCCC")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if c_idx == 6:
                cell.fill = PatternFill("solid", start_color=STATUS_COLORS_EXCEL.get(status, bg))
                cell.font = Font(name="Arial", size=10, bold=True)
            else:
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(name="Arial", size=10)
    widths = [22, 22, 25, 12, 14, 14, 40, 12, 10]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    wb.save(path)
